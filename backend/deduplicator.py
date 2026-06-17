


#!/usr/bin/env python3
"""
Enhanced PDF/File Deduplication Engine
Supports in-memory processing for web API integration.
"""
import os
import io
import hashlib
import shutil
import zipfile
import tempfile
from collections import defaultdict
from typing import Optional, List, Dict, Any
 
 
def calculate_sha256(data: bytes) -> str:
    """Calculate the SHA-256 hash of a bytes object."""
    return hashlib.sha256(data).hexdigest()
 
 
def calculate_sha256_file(file_path: str, chunk_size: int = 65536) -> Optional[str]:
    """Calculate the SHA-256 hash of a file on disk in chunks."""
    sha256 = hashlib.sha256()
    try:
        with open(file_path, "rb") as f:
            while chunk := f.read(chunk_size):
                sha256.update(chunk)
        return sha256.hexdigest()
    except OSError:
        return None
 
 
def are_files_identical(file1: str, file2: str, chunk_size: int = 65536) -> bool:
    """Compare two files byte-by-byte in chunks to guarantee they are identical."""
    try:
        with open(file1, "rb") as f1, open(file2, "rb") as f2:
            while True:
                chunk1 = f1.read(chunk_size)
                chunk2 = f2.read(chunk_size)
                if chunk1 != chunk2:
                    return False
                if not chunk1:
                    return True
    except OSError:
        return False
 
 
def get_file_category(filename: str) -> str:
    """Determine a human-readable file category based on the extension."""
    ext = os.path.splitext(filename)[1].lower()
    categories = {
        ".pdf": "PDF Document",
        ".txt": "Plain Text",
        ".csv": "CSV Spreadsheet",
        ".xlsx": "Excel Spreadsheet",
        ".xls": "Excel Spreadsheet",
        ".docx": "Word Document",
        ".doc": "Word Document",
        ".png": "Image",
        ".jpg": "Image",
        ".jpeg": "Image",
        ".gif": "Image",
        ".webp": "Image",
        ".zip": "Archive",
        ".rar": "Archive",
        ".tar": "Archive",
        ".gz": "Archive",
        ".7z": "Archive",
        ".mp3": "Audio",
        ".wav": "Audio",
        ".mp4": "Video",
        ".mkv": "Video",
        ".mov": "Video",
    }
    return categories.get(ext, f"{ext[1:].upper() if ext else 'Unknown'} File")
 
 
def get_unique_zip_dest_path(original_path: str, allocated_paths: set) -> str:
    """
    Finds a unique flat path in ZIP for original_path to avoid collisions.
    Appends _dup1, _dup2, etc., if a collision is found.
    Tracks path allocations in allocated_paths.
    """
    base_name = os.path.basename(original_path)
    name, ext = os.path.splitext(base_name)
    dest_path = base_name
    counter = 1
    while dest_path in allocated_paths:
        dest_path = f"{name}_dup{counter}{ext}"
        counter += 1
    allocated_paths.add(dest_path)
    return dest_path
 
 
def generate_excel_report_bytes(duplicate_mappings: List[tuple]) -> bytes:
    """Writes duplicate mappings to an Excel workbook in-memory and returns bytes."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duplicate Report"
 
    # Define headers
    headers = [
        "Duplicate File Name",
        "Original Path",
        "New Path",
        "Matched Original Path",
        "Category"
    ]
    
    # Stylize header row (dark slate blue background with white text)
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
 
    ws.row_dimensions[1].height = 25
 
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
 
    # Add data rows
    font_data = Font(name="Calibri", size=11)
    align_left = Alignment(horizontal="left", vertical="center")
    
    for row_idx, (dup_name, orig_path, new_path, match_path, cat) in enumerate(duplicate_mappings, 2):
        ws.row_dimensions[row_idx].height = 20
        row_data = [dup_name, orig_path, new_path, match_path, cat]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.alignment = align_left
 
    # Auto-fit column widths based on content length
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
 
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
 
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
 
 
def deduplicate_zip_bytes(zip_bytes: bytes, recursive: bool = True) -> dict:
    """
    Process a ZIP file (as bytes) and deduplicate its contents using size sorting,
    lazy hash computation, and byte-by-byte content comparison.
    """
    try:
        input_zip = zipfile.ZipFile(io.BytesIO(zip_bytes), "r")
    except zipfile.BadZipFile:
        return {"error": "Invalid ZIP file"}
 
    # Read all files from the zip
    all_entries = []
    for info in input_zip.infolist():
        if info.is_dir():
            continue
 
        name = info.filename
        basename = os.path.basename(name)
        
        # Ignore macOS system files and folders
        if basename.lower() == ".ds_store" or basename.lower().startswith("._"):
            continue
        if "__macosx/" in name.lower() or name.lower().startswith("__macosx/"):
            continue
 
        try:
            data = input_zip.read(name)
        except Exception:
            continue
 
        # For non-recursive: skip files in subdirectories
        if not recursive and "/" in name.rstrip("/"):
            parts = name.split("/")
            if len(parts) > 2 or (len(parts) == 2 and parts[1] != ""):
                continue
 
        all_entries.append({
            "name": name,
            "basename": os.path.basename(name),
            "data": data,
            "size": len(data),
            "category": get_file_category(name),
            "hash": None,  # Lazily computed
        })
 
    input_zip.close()
 
    # Track unique files by size
    unique_by_size = defaultdict(list)
    unique_files = []
    duplicate_files = []
    allocated_paths = set()
    duplicate_mappings = []
 
    for entry in all_entries:
        size = entry["size"]
        is_duplicate = False
        duplicate_of = None
 
        if size in unique_by_size:
            # Lazily calculate hash of current entry
            hash_current = calculate_sha256(entry["data"])
            
            for candidate in unique_by_size[size]:
                # Lazily compute hash for candidate if not done yet
                if candidate["hash"] is None:
                    candidate["hash"] = calculate_sha256(candidate["data"])
                
                # Compare hashes, then byte-by-byte content
                if candidate["hash"] == hash_current:
                    if candidate["data"] == entry["data"]:
                        is_duplicate = True
                        duplicate_of = candidate
                        break
            
            if not is_duplicate:
                entry["hash"] = hash_current
                unique_by_size[size].append(entry)
                unique_files.append(entry)
        else:
            # First file of this size. Assume unique (hash is None)
            unique_by_size[size].append(entry)
            unique_files.append(entry)
 
        if is_duplicate:
            resolved_dup_name = get_unique_zip_dest_path(entry["name"], allocated_paths)
            duplicate_files.append({
                **entry,
                "resolved_name": resolved_dup_name,
                "duplicate_of": duplicate_of["name"],
                "duplicate_of_basename": duplicate_of["basename"],
            })
            duplicate_mappings.append((
                resolved_dup_name,
                entry["name"],
                resolved_dup_name,
                duplicate_of["name"],
                entry["category"]
            ))
 
    stats = {
        "scanned": len(all_entries),
        "unique": len(unique_files),
        "duplicates": len(duplicate_files),
        "bytes_saved": sum(d["size"] for d in duplicate_files),
    }
 
    # Build result ZIP (unique files only, keeping original paths)
    result_buf = io.BytesIO()
    with zipfile.ZipFile(result_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for f in unique_files:
            zout.writestr(f["name"], f["data"])
    result_zip = result_buf.getvalue()
 
    # Build duplicates ZIP (flat structure with resolved names and Excel report)
    dups_buf = io.BytesIO()
    with zipfile.ZipFile(dups_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for f in duplicate_files:
            zout.writestr(f["resolved_name"], f["data"])
        
        # Add Excel report if there are duplicate files
        if len(duplicate_mappings) > 0:
            try:
                report_bytes = generate_excel_report_bytes(duplicate_mappings)
                zout.writestr("duplicates_report.xlsx", report_bytes)
            except Exception:
                pass
    duplicates_zip = dups_buf.getvalue()
 
    return {
        "unique_files": [
            {"name": f["name"], "basename": f["basename"], "size": f["size"], "category": f["category"]}
            for f in unique_files
        ],
        "duplicate_files": [
            {
                "name": f["name"],
                "basename": f["resolved_name"],  # resolved name
                "size": f["size"],
                "category": f["category"],
                "duplicate_of": f["duplicate_of"],
                "duplicate_of_basename": f["duplicate_of_basename"],
            }
            for f in duplicate_files
        ],
        "stats": stats,
        "result_zip": result_zip,
        "duplicates_zip": duplicates_zip,
    }
 
 
def deduplicate_directory(source_dir: str, recursive: bool = True) -> dict:
    """
    Deduplicate files in a local directory.
    Returns the same structure as deduplicate_zip_bytes.
    """
    source_dir_abs = os.path.abspath(source_dir)
    duplicates_dir_abs = os.path.join(source_dir_abs, "duplicates")
    all_entries = []
 
    # Helper to load file bytes
    def try_read_file(fpath):
        try:
            with open(fpath, "rb") as f:
                return f.read()
        except Exception:
            return None
 
    if recursive:
        for root, dirs, files in os.walk(source_dir_abs):
            root_abs = os.path.abspath(root)
            if root_abs.startswith(duplicates_dir_abs):
                continue
            dirs[:] = [d for d in dirs if os.path.join(root_abs, d) != duplicates_dir_abs]
            
            for fname in files:
                # Skip macOS system files and folders
                if fname.lower() == ".ds_store" or fname.lower().startswith("._"):
                    continue
                fpath = os.path.join(root, fname)
                if os.path.abspath(fpath).startswith(duplicates_dir_abs):
                    continue
                if "__macosx/" in fpath.lower() or "/__macosx/" in fpath.lower():
                    continue
                data = try_read_file(fpath)
                if data is None:
                    continue
                rel = os.path.relpath(fpath, source_dir_abs)
                all_entries.append({
                    "name": rel,
                    "basename": fname,
                    "data": data,
                    "size": len(data),
                    "category": get_file_category(fname),
                    "hash": None,
                })
    else:
        try:
            for entry in os.scandir(source_dir_abs):
                if entry.is_file():
                    # Skip macOS system files and folders
                    if entry.name.lower() == ".ds_store" or entry.name.lower().startswith("._"):
                        continue
                    if os.path.abspath(entry.path).startswith(duplicates_dir_abs):
                        continue
                    if "__macosx/" in entry.path.lower() or "/__macosx/" in entry.path.lower():
                        continue
                    data = try_read_file(entry.path)
                    if data is None:
                        continue
                    all_entries.append({
                        "name": entry.name,
                        "basename": entry.name,
                        "data": data,
                        "size": len(data),
                        "category": get_file_category(entry.name),
                        "hash": None,
                    })
        except OSError:
            pass
 
    # Track unique files by size
    unique_by_size = defaultdict(list)
    unique_files = []
    duplicate_files = []
    allocated_paths = set()
    duplicate_mappings = []
 
    for entry in all_entries:
        size = entry["size"]
        is_duplicate = False
        duplicate_of = None
 
        if size in unique_by_size:
            # Lazily calculate hash of current entry
            hash_current = calculate_sha256(entry["data"])
            
            for candidate in unique_by_size[size]:
                # Lazily compute hash for candidate if not done yet
                if candidate["hash"] is None:
                    candidate["hash"] = calculate_sha256(candidate["data"])
                
                # Compare hashes, then byte-by-byte content
                if candidate["hash"] == hash_current:
                    if candidate["data"] == entry["data"]:
                        is_duplicate = True
                        duplicate_of = candidate
                        break
            
            if not is_duplicate:
                entry["hash"] = hash_current
                unique_by_size[size].append(entry)
                unique_files.append(entry)
        else:
            # First file of this size. Assume unique (hash is None)
            unique_by_size[size].append(entry)
            unique_files.append(entry)
 
        if is_duplicate:
            resolved_dup_name = get_unique_zip_dest_path(entry["name"], allocated_paths)
            duplicate_files.append({
                **entry,
                "resolved_name": resolved_dup_name,
                "duplicate_of": duplicate_of["name"],
                "duplicate_of_basename": duplicate_of["basename"],
            })
            duplicate_mappings.append((
                resolved_dup_name,
                entry["name"],
                resolved_dup_name,
                duplicate_of["name"],
                entry["category"]
            ))
 
    stats = {
        "scanned": len(all_entries),
        "unique": len(unique_files),
        "duplicates": len(duplicate_files),
        "bytes_saved": sum(d["size"] for d in duplicate_files),
    }
 
    # Build result ZIP (unique files only, keeping original paths)
    result_buf = io.BytesIO()
    with zipfile.ZipFile(result_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for f in unique_files:
            zout.writestr(f["name"], f["data"])
    result_zip = result_buf.getvalue()
 
    # Build duplicates ZIP (flat structure with resolved names and Excel report)
    dups_buf = io.BytesIO()
    with zipfile.ZipFile(dups_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for f in duplicate_files:
            zout.writestr(f["resolved_name"], f["data"])
        
        # Add Excel report if there are duplicate files
        if len(duplicate_mappings) > 0:
            try:
                report_bytes = generate_excel_report_bytes(duplicate_mappings)
                zout.writestr("duplicates_report.xlsx", report_bytes)
            except Exception:
                pass
    duplicates_zip = dups_buf.getvalue()
 
    return {
        "unique_files": [
            {"name": f["name"], "basename": f["basename"], "size": f["size"], "category": f["category"]}
            for f in unique_files
        ],
        "duplicate_files": [
            {
                "name": f["name"],
                "basename": f["resolved_name"],
                "size": f["size"],
                "category": f["category"],
                "duplicate_of": f["duplicate_of"],
                "duplicate_of_basename": f["duplicate_of_basename"],
            }
            for f in duplicate_files
        ],
        "stats": stats,
        "result_zip": result_zip,
        "duplicates_zip": duplicates_zip,
    }